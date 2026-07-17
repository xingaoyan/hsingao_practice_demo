#!/usr/bin/env python3
"""
Excel 解析器
"""

from pathlib import Path
from typing import Dict, List, Optional, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseParser, ParseResult, PageInfo


class ExcelParser(BaseParser):
    """Excel 解析器"""
    
    def parse(self, file_path: Path, document_id: str) -> ParseResult:
        """解析 Excel 文件"""
        self.clear_warnings()
        
        try:
            workbook = load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as e:
            self.add_warning(f"无法打开 Excel 文件: {e}")
            return self._create_empty_result(file_path, document_id)
        
        pages = []
        metadata = self._extract_metadata(workbook, file_path)
        
        # 提取每个工作表
        for sheet_num, sheet_name in enumerate(workbook.sheetnames, 1):
            sheet = workbook[sheet_name]
            page_info = self._parse_sheet(sheet, sheet_name, sheet_num)
            pages.append(page_info)
        
        # 提取标题
        title = self._extract_title(workbook, file_path)
        
        workbook.close()
        
        return ParseResult(
            document_id=document_id,
            title=title,
            source_path=str(file_path),
            file_type="xlsx",
            pages=pages,
            metadata=metadata,
            warnings=self.warnings
        )
    
    def _parse_sheet(self, sheet: Worksheet, sheet_name: str, sheet_number: int) -> PageInfo:
        """解析单个工作表"""
        try:
            # 提取表头
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value is not None:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append("")
            
            # 提取数据行
            data_rows = []
            for row in range(2, sheet.max_row + 1):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value).strip())
                    else:
                        row_data.append("")
                
                # 跳过空行
                if any(cell for cell in row_data):
                    data_rows.append(row_data)
            
            # 创建表格文本
            table_text = self._create_table_text(headers, data_rows)
            
            # 提取合并单元格信息
            merged_cells = self._extract_merged_cells(sheet)
            
            return PageInfo(
                page_number=sheet_number,
                text=table_text,
                images=[],
                tables=[{
                    "sheet_name": sheet_name,
                    "headers": headers,
                    "row_count": len(data_rows),
                    "column_count": len(headers),
                    "merged_cells": merged_cells
                }]
            )
            
        except Exception as e:
            self.add_warning(f"解析工作表 {sheet_name} 时出错: {e}")
            return PageInfo(
                page_number=sheet_number,
                text="",
                images=[],
                tables=[]
            )
    
    def _create_table_text(self, headers: List[str], data_rows: List[List[str]]) -> str:
        """创建表格文本"""
        if not headers:
            return ""
        
        # 创建 Markdown 表格
        table_lines = []
        table_lines.append("| " + " | ".join(headers) + " |")
        table_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        
        for row in data_rows:
            # 确保行长度与表头一致
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
            elif len(row) > len(headers):
                row = row[:len(headers)]
            
            table_lines.append("| " + " | ".join(row) + " |")
        
        return "\n".join(table_lines)
    
    def _extract_merged_cells(self, sheet: Worksheet) -> List[Dict]:
        """提取合并单元格信息"""
        merged_cells = []
        try:
            for merged_range in sheet.merged_cells.ranges:
                merged_cells.append({
                    "range": str(merged_range),
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col
                })
        except Exception as e:
            self.add_warning(f"提取合并单元格信息时出错: {e}")
        return merged_cells
    
    def _extract_metadata(self, workbook, file_path: Path) -> Dict[str, Any]:
        """提取工作簿元数据"""
        metadata = self.get_metadata(file_path)
        
        try:
            # 提取核心属性
            properties = workbook.properties
            if properties:
                metadata.update({
                    "title": properties.title or "",
                    "creator": properties.creator or "",
                    "subject": properties.subject or "",
                    "description": properties.description or "",
                    "keywords": properties.keywords or "",
                    "category": properties.category or "",
                    "created": str(properties.created) if properties.created else "",
                    "modified": str(properties.modified) if properties.modified else "",
                    "last_modified_by": properties.lastModifiedBy or ""
                })
            
            metadata["sheet_count"] = len(workbook.sheetnames)
            metadata["sheet_names"] = workbook.sheetnames
            
        except Exception as e:
            self.add_warning(f"提取元数据时出错: {e}")
        
        return metadata
    
    def _extract_title(self, workbook, file_path: Path) -> str:
        """提取工作簿标题"""
        # 首先尝试从元数据获取
        try:
            if workbook.properties.title:
                return workbook.properties.title
        except:
            pass
        
        # 使用第一个工作表名称作为标题
        try:
            if workbook.sheetnames:
                return workbook.sheetnames[0]
        except:
            pass
        
        # 使用文件名作为标题
        return file_path.stem
    
    def _create_empty_result(self, file_path: Path, document_id: str) -> ParseResult:
        """创建空结果"""
        return ParseResult(
            document_id=document_id,
            title=file_path.stem,
            source_path=str(file_path),
            file_type="xlsx",
            pages=[],
            metadata=self.get_metadata(file_path),
            warnings=self.warnings
        )